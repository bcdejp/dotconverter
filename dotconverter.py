# -*- coding: utf-8 -*-
#
#
#

import cv2
import numpy as np
import openpyxl

# 変換後のイメージ設定
width    = 32 # 幅
height   = 32 # 高さ
colornum = 15 # 色数

# パレット設定
step_hue = 30 # 色相段数
step_stc = 15 # 彩度段数
step_vlb = 15 # 明度段数

# 減色処理
def sub_color(src, K):
    # 次元数を1落とす
    Z = src.reshape((-1,3))
    # float32型に変換
    Z = np.float32(Z)
    # 基準の定義
    criteria = (cv2.TERM_CRITERIA_EPS + cv2.TERM_CRITERIA_MAX_ITER, 10, 1.0)
    # K-means法で減色
    ret, label, center = cv2.kmeans(Z, K, None, criteria, 10, cv2.KMEANS_RANDOM_CENTERS)
    # UINT8に変換
    center = np.uint8(center)
    res = center[label.flatten()]
    # 配列の次元数と入力画像と同じに戻す
    return res.reshape((src.shape))

# モザイク処理
def mosaic(img, width, height):
    # 画像の高さ、幅、チャンネル数
    h, w, ch = img.shape
    # 縮小→拡大でモザイク加工
    img = cv2.resize(img,(int(width), int(height)))
    return img

# ドット絵化
def pixel_art(img, width, height, K):
    # モザイク処理
    img = mosaic(img, width, height)
    # 減色処理
    img = sub_color(img, K)
    return img

# ラベル化
def num3label(x, y, z):
    label = str(x)+'+'+str(y)+'+'+str(z)
    return label

# パレット変換(HSV)
def hsv2palette(x, y, z):
    pallette = [int(x/179*step_hue), int(y/255*step_stc), int(z/255*step_vlb)]
    return pallette

# ラベル→数値
def lable_num3(label):
    return label.split('+')


# メイン処理
if __name__ == '__main__':
    # 入力画像を取得
    img = cv2.imread("./raw.png")
    # ドット絵化(縮小＋原色)
    img = pixel_art(img, width, height, colornum)
    # ドット絵をファイル出力
    cv2.imwrite("./dot.png", img)

    # BGRのテーブル(Excelファイル出力時に使用)
    bgr_array = np.asarray(img) #numpyで扱える配列をつくる

    # 色配列の変換(BGR->HSV)
    # BGR(256,256,256)->HSV(180,256,256)
    hsv_img = cv2.cvtColor(img, cv2.COLOR_BGR2HSV)

    # パレットカラーに変換(HSV)
    hsv_array = np.asarray(hsv_img) #numpyで扱える配列をつくる
    for x in range(0, width):
        for y in range(0, height):
            plt = hsv2palette(int(hsv_array[x,y,0]),int(hsv_array[x,y,1]),int(hsv_array[x,y,2]))
            hsv_array[x,y,] = plt

    # 色(HSV)カウント(多く使われている色から降順ソート)
    dict_color = {}
    for x in range(0, width):
        for y in range(0, height):
            #print(hsv_array[x,y,])
            color = num3label(int(hsv_array[x,y,0]),int(hsv_array[x,y,1]),int(hsv_array[x,y,2]))
            if color not in dict_color:
                dict_color[color] = 0
            else:
                dict_color[color] = dict_color[color] + 1
    list_color = sorted(dict_color.items(), key=lambda x:-x[1])
    #print(dict_color)

    # カラーインディクスを作成
    # 色数が多い順からインディクスをふる
    # インディクスは1始まりとする
    dict_index = dict_color.copy()
    counter = 1
    for item in list_color:
        dict_index[item[0]]=counter
        counter = counter + 1

    # ドットテーブルの宣言
    dot_table = [[0 for i in range(width)] for j in range(height)]

    # ドットテーブルにインディクスを付与
    hsv_array = np.asarray(hsv_img) #numpyで扱える配列をつくる
    for x in range(0, width):
        for y in range(0, height):
            color = num3label(int(hsv_array[x,y,0]),int(hsv_array[x,y,1]),int(hsv_array[x,y,2]))
            dot_table[x][y] = dict_index[color]
    
    # テスト表示(dot)       
    for x in range(0, width):
        for y in range(0, height):
            print("{0:2d}".format(dot_table[x][y]), end='')
        print("")
    
    # パレットを表示
    counter = 1
    for item in list_color:
        print(str(counter)+ ":" + str(item[1]) + ":" + item[0])
        counter = counter + 1
    
    # Excelファイル出力
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'my_design'

    # ドットイメージ出力
    start_row    = 1
    start_column = 1
    for x in range(0, width):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(x+1)].width = 3
        for y in range(0, height):
            # colum,row
            column = start_column + y  
            row    = start_row    + x
            # セルindex
            sheet.cell(row, column).value = dot_table[x][y]
            # セルcolor
            hex_color = '%02X%02X%02X' % (int(bgr_array[x,y,2]),int(bgr_array[x,y,1]),int(bgr_array[x,y,0]))
            fill = openpyxl.styles.PatternFill(patternType='solid', fgColor=hex_color)
            sheet.cell(row, column).fill = fill
    
    # パレット出力
    start_row    = width + 1
    start_column = 1
    sheet.cell(start_column, start_row).value   = "No"       # No
    sheet.cell(start_column, start_row+1).value = "色相"     # 色相
    sheet.cell(start_column, start_row+2).value = "鮮やかさ" # 鮮やかさ
    sheet.cell(start_column, start_row+3).value = "明るさ"   # 明るさ  
    start_column = start_column + 1 
    counter = 0
    for item in list_color:
        sheet.cell(start_column+counter, start_row).value  = str(counter+1)         # No
        label_array = lable_num3(item[0])
        sheet.cell(start_column+counter, start_row+1).value = str(label_array[0]) # 色相
        sheet.cell(start_column+counter, start_row+2).value = str(label_array[1]) # 鮮やかさ
        sheet.cell(start_column+counter, start_row+3).value = str(label_array[2]) # 明るさ
        counter = counter + 1

    wb.save('my_design.xlsx')
    
    
